import random
import io
import math
import streamlit as st
from openpyxl import Workbook
import pandas as pd

# ── 页面配置 ──────────────────────────────────────────────
st.set_page_config(page_title="随机数生成器", page_icon="🎲", layout="wide")
st.title("🎲 随机数生成器")
st.caption("自由配置分段、趋势和波动幅度，生成随机数并导出")

# ── 初始化 session_state ──────────────────────────────────
if "segments" not in st.session_state:
    st.session_state.segments = [
        {"count": 1, "trend": "平稳", "start": 0.0, "end": 0.0},
    ]
if "gen_counter" not in st.session_state:
    st.session_state.gen_counter = 0

# ── 全局参数 ──────────────────────────────────────────────
st.subheader("全局参数")

volatility = 0.05
if "use_gauss" not in st.session_state:
    st.session_state.use_gauss = "均匀分布"

col1, col2 = st.columns(2)

with col1:
    decimals = st.number_input("小数位数", min_value=0, max_value=10, value=5, step=1)

with col2:
    st.radio(
        "分布模式（平稳趋势时生效）",
        options=["均匀分布", "高斯分布", "三角分布", "指数分布（右偏）", "指数分布（左偏）"],
        index=0,
        horizontal=False,
        key="use_gauss",
        help="高斯分布：数值集中在中心附近\n均匀分布：数值在范围内均匀分散\n三角分布：中心概率高，两侧递减\n指数分布：一端概率高，向另一端递减"
    )

use_gauss = st.session_state.use_gauss in ["高斯分布", "三角分布"]

if use_gauss:
    volatility = st.slider("波动幅度（标准差）", min_value=0.0, max_value=10.0, value=0.5, step=0.01, format="%.3f")
else:
    volatility = 0.5  # 均匀/指数分布不需要标准差

# ── 分段设置 ──────────────────────────────────────────────
st.subheader("分段设置")


def add_segment():
    st.session_state.segments.append({"count": 10, "trend": "平稳", "start": 1.0, "end": 1.0})


def remove_segment(idx):
    if len(st.session_state.segments) > 1:
        st.session_state.segments.pop(idx)


for i, seg in enumerate(st.session_state.segments):
    with st.container():
        cols = st.columns([1.5, 1.5, 1.5, 1.5, 0.6])
        with cols[0]:
            st.session_state.segments[i]["count"] = st.number_input(
                f"第{i+1}段 数量", min_value=1, max_value=9999, value=seg["count"], step=1, key=f"count_{i}"
            )
        with cols[1]:
            trend_options = ["平稳", "上升", "下降"]
            idx = trend_options.index(seg["trend"]) if seg["trend"] in trend_options else 0
            st.session_state.segments[i]["trend"] = st.selectbox(
                f"第{i+1}段 趋势", trend_options, index=idx, key=f"trend_{i}"
            )
        with cols[2]:
            st.session_state.segments[i]["start"] = st.number_input(
                f"第{i+1}段 起始值", value=seg["start"], step=0.1, format="%.5f", key=f"start_{i}"
            )
        with cols[3]:
            st.session_state.segments[i]["end"] = st.number_input(
                f"第{i+1}段 终止值", value=seg["end"], step=0.1, format="%.5f", key=f"end_{i}"
            )
        with cols[4]:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            st.button("❌", key=f"del_{i}", on_click=remove_segment, args=(i,))

st.button("＋ 添加分段", on_click=add_segment)

# ── 生成随机数 ────────────────────────────────────────────


def generate_numbers(segments, decimals, volatility, dist_mode):
    result = []
    for seg in segments:
        count = seg["count"]
        trend = seg["trend"]
        start = seg["start"]
        end = seg["end"]

        low = min(start, end)
        high = max(start, end)

        for i in range(count):
            if trend == "平稳":
                center = (start + end) / 2
            else:
                ratio = i / max(count - 1, 1)
                center = start + (end - start) * ratio

            val = _generate_one(center, low, high, volatility, dist_mode, decimals)
            result.append(val)
    return result


def _generate_one(center, low, high, volatility, dist_mode, decimals):
    """根据分布模式生成一个随机数，严格在 [low, high] 范围内"""
    max_attempts = 100

    if dist_mode == "均匀分布":
        return round(random.uniform(low, high), decimals)

    elif dist_mode == "三角分布":
        # triangular: 中心概率最高，向两边线性递减
        for _ in range(max_attempts):
            val = round(random.triangular(low, high, center), decimals)
            if low <= val <= high:
                return val
        return round(center, decimals)

    elif dist_mode == "指数分布（右偏）":
        # 左边概率高，右边递减（lambda 越大越集中在左侧）
        # 映射：center -> low 方向概率高
        rate = 1 / max(volatility, 0.001)
        for _ in range(max_attempts):
            val = round(low + random.expovariate(rate), decimals)
            if low <= val <= high:
                return val
        return round(random.triangular(low, center, high), decimals)

    elif dist_mode == "指数分布（左偏）":
        # 右边概率高，左边递减
        rate = 1 / max(volatility, 0.001)
        for _ in range(max_attempts):
            val = round(high - random.expovariate(rate), decimals)
            if low <= val <= high:
                return val
        return round(random.triangular(low, high, center), decimals)

    else:
        # 高斯分布（默认）
        for _ in range(max_attempts):
            val = round(random.gauss(center, volatility), decimals)
            if low <= val <= high:
                return val
        return max(low, min(high, val))


def generate_excel(numbers):
    wb = Workbook()
    ws = wb.active
    ws.title = "随机数"
    ws["A1"] = "序号"
    ws["B1"] = "数值"
    for i, num in enumerate(numbers, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=num)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 操作按钮 ──────────────────────────────────────────────
st.markdown("---")
btn_col1, btn_col2, btn_col3 = st.columns(3)

with btn_col1:
    gen_clicked = st.button("🎲 生成随机数", type="primary", use_container_width=True)
with btn_col2:
    copy_clicked = st.button("📋 一键复制", use_container_width=True)
with btn_col3:
    download_placeholder = st.empty()

# ── 生成逻辑 ──────────────────────────────────────────────
if gen_clicked:
    numbers = generate_numbers(st.session_state.segments, decimals, volatility,
                                  st.session_state.use_gauss)
    st.session_state.numbers = numbers
    st.session_state.gen_counter += 1

if "numbers" in st.session_state and st.session_state.numbers:
    numbers = st.session_state.numbers

    # 折线图预览（用 DataFrame 确保横纵坐标正确）
    st.subheader("📊 预览")
    df = pd.DataFrame({"序号": range(1, len(numbers) + 1), "数值": numbers})
    st.line_chart(df, x="序号", y="数值", height=350)

    # 数据展示（用 text_area 方便全选复制）
    st.subheader("📋 数据")
    display_text = "  ".join(str(n) for n in numbers)
    st.text_area("随机数序列（可直接选中复制）", value=display_text, height=150, key=f"data_display_{st.session_state.gen_counter}")

    # 一键复制：通过 JS 找到页面上的 textarea 并复制其内容
    if copy_clicked:
        js_copy = """
        <script>
        function doCopy() {
            // 找到 Streamlit 页面中所有 textarea，取最后一个（就是数据展示区）
            var textareas = parent.document.querySelectorAll('textarea');
            if (textareas.length > 0) {
                var ta = textareas[textareas.length - 1];
                ta.select();
                ta.setSelectionRange(0, ta.value.length);
                try {
                    parent.document.execCommand('copy');
                    document.getElementById('copy-msg').innerText = '✅ 已复制到剪贴板！';
                } catch(e) {
                    document.getElementById('copy-msg').innerText = '❌ 复制失败，请手动 Ctrl+A 全选后复制';
                }
            } else {
                document.getElementById('copy-msg').innerText = '❌ 未找到数据区域';
            }
            setTimeout(() => {
                var el = document.getElementById('copy-msg');
                if (el) el.innerText = '';
            }, 3000);
        }
        doCopy();
        </script>
        <p id="copy-msg" style="color:green;font-weight:bold;font-size:14px;"></p>
        """
        st.components.v1.html(js_copy, height=30)

    # Excel 下载
    excel_buf = generate_excel(numbers)
    with download_placeholder:
        st.download_button(
            label="📥 下载 Excel",
            data=excel_buf,
            file_name="随机数.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
